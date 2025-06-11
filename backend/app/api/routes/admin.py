from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, String, text
from typing import List, Optional
import logging

from app.database import get_db
from app.models.user import User
from app.models.queue import QueueEntry, QueueStatus
from app.models.video import VideoSettings
from app.schemas.queue import QueueResponse
from app.schemas.video import VideoSettingsResponse, VideoSettingsUpdate
from app.schemas import AdminUserCreate, UserResponse, UserUpdate
from app.security import get_admin_user
from app.services.user import create_user
from app.services.queue import get_all_queue_entries
from app.services.archive import get_archive_statistics, cleanup_old_completed_entries, archive_queue_entry
from app.models.archive import ArchivedQueueEntry
from fastapi.responses import StreamingResponse
import io
import csv
from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()
logger = logging.getLogger(__name__)

# Полный словарь для маппинга названий программ на коды
PROGRAM_MAPPING = {
    # БАКАЛАВРИАТ
    # Русские названия
    "бухгалтерский учёт": "accounting",
    "бухгалтерский учет": "accounting",
    "прикладная лингвистика": "appliedLinguistics",
    "экономика и наука о данных": "economicsDataScience",
    "финансы": "finance",
    "гостеприимство": "hospitality",
    "международная журналистика": "internationalJournalism",
    "международное право": "internationalLaw",
    "международные отношения": "internationalRelations",
    "it": "it",
    "ит": "it",
    "юриспруденция": "jurisprudence",
    "менеджмент": "management",
    "маркетинг": "marketing",
    "психология": "psychology",
    "туризм": "tourism",
    "переводческое дело": "translation",
    
    # Казахские названия
    "бухгалтерлік есеп": "accounting",
    "қолданбалы лингвистика": "appliedLinguistics",
    "экономика және деректер ғылымы": "economicsDataScience",
    "қаржы": "finance",
    "қонақжайлылық": "hospitality",
    "халықаралық журналистика": "internationalJournalism",
    "халықаралық құқық": "internationalLaw",
    "халықаралық қатынастар": "internationalRelations",
    "құқықтану": "jurisprudence",
    "аударма ісі": "translation",
    
    # Английские названия
    "accounting": "accounting",
    "applied linguistics": "appliedLinguistics",
    "economics and data science": "economicsDataScience",
    "finance": "finance",
    "hospitality": "hospitality",
    "international journalism": "internationalJournalism",
    "international law": "internationalLaw",
    "international relations": "internationalRelations",
    "law": "jurisprudence",
    "management": "management",
    "marketing": "marketing",
    "psychology": "psychology",
    "tourism": "tourism",
    "translation studies": "translation",
    
    # МАГИСТРАТУРА
    # Русские названия
    "политология и международные отношения": "politicalInternationalRelations",
    "конкурентное право": "competitionLaw",
    "консультативная психология": "consultingPsychology",
    "экономика": "economics",
    "право интеллектуальной собственности и бизнеса": "intellectualPropertyLaw",
    "право it": "itLaw",
    "право ит": "itLaw",
    
    # Казахские названия
    "саясаттану және халықаралық қатынастар": "politicalInternationalRelations",
    "бәсекелестік құқық": "competitionLaw",
    "консультативті психология": "consultingPsychology",
    "зияткерлік меншік және бизнес құқық": "intellectualPropertyLaw",
    "құқық it": "itLaw",
    
    # Английские названия
    "political science and international relations": "politicalInternationalRelations",
    "competition law": "competitionLaw",
    "counselling psychology": "consultingPsychology",
    "economics": "economics",
    "intellectual property and business law": "intellectualPropertyLaw",
    "it law": "itLaw",
    
    # ДОКТОРАНТУРА
    # Русские названия
    "право": "law",
    "phd по экономике": "phdEconomics",
    
    # Казахские названия
    "құқық": "law",
    "экономика саласындағы phd": "phdEconomics",
    
    # Английские названия
    "phd in law": "law",
    "phd in economics": "phdEconomics"
}

@router.post("/sync/google-sheets/full")
def full_sync_to_google_sheets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Полная синхронизация всех данных из архива в Google Sheets"""
    try:
        from app.services.google_sheets import google_sheets_service
        
        logger.info("🚀 Запуск полной синхронизации с Google Sheets...")
        
        result = google_sheets_service.sync_all_data(db)
        
        if result.get("success"):
            logger.info(f"✅ Полная синхронизация завершена успешно")
            return {
                "success": True,
                "message": "Full synchronization completed successfully",
                "details": result
            }
        else:
            logger.error(f"❌ Ошибка полной синхронизации: {result.get('error')}")
            return {
                "success": False,
                "message": "Full synchronization failed",
                "error": result.get("error")
            }
            
    except Exception as e:
        logger.error(f"❌ Исключение при полной синхронизации: {e}")
        raise HTTPException(status_code=500, detail=f"Full sync failed: {str(e)}")

@router.get("/sync/test-now")
def test_sync_now(db: Session = Depends(get_db)):
    """Временный роут для быстрого тестирования синхронизации"""
    try:
        from app.services.google_sheets import google_sheets_service
        result = google_sheets_service.sync_all_data(db)
        return result
    except Exception as e:
        return {"error": str(e)}

@router.get("/sync/google-sheets/status")
def get_sync_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Получить статус синхронизации с Google Sheets"""
    try:
        from app.services.scheduler import realtime_sync
        from app.services.google_sheets import google_sheets_service
        
        # Проверяем доступность Google Sheets API
        is_available = google_sheets_service._is_available()
        
        # Получаем статистику из архива
        total_archive_entries = db.query(ArchivedQueueEntry).count()
        
        # Пытаемся получить количество строк в Google Sheets
        sheets_rows = 0
        sheets_error = None
        
        if is_available:
            try:
                # Получаем данные из Google Sheets для подсчета
                range_name = f'{google_sheets_service.sheet_name}!A:A'
                result = google_sheets_service.service.spreadsheets().values().get(
                    spreadsheetId=google_sheets_service.spreadsheet_id,
                    range=range_name
                ).execute()
                
                values = result.get('values', [])
                sheets_rows = len(values) - 1 if values else 0  # -1 для заголовка
                
            except Exception as e:
                sheets_error = str(e)
        
        # Получаем статистику синхронизации
        sync_stats = realtime_sync.get_sync_stats()
        
        return {
            "success": True,
            "google_sheets_available": is_available,
            "total_archive_entries": total_archive_entries,
            "google_sheets_rows": sheets_rows,
            "sheets_error": sheets_error,
            "sync_stats": sync_stats,
            "needs_full_sync": sheets_rows < total_archive_entries
        }
        
    except Exception as e:
        logger.error(f"❌ Ошибка получения статуса синхронизации: {e}")
        raise HTTPException(status_code=500, detail=f"Status check failed: {str(e)}")

@router.post("/sync/google-sheets/test")
def test_google_sheets_connection(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Тестирование подключения к Google Sheets"""
    try:
        from app.services.google_sheets import google_sheets_service
        
        if not google_sheets_service._is_available():
            return {
                "success": False,
                "message": "Google Sheets API недоступен"
            }
        
        # Пробуем записать тестовые данные
        test_data = [["Test", "Data", "Connection", datetime.now().isoformat()]]
        
        try:
            # Добавляем тестовую строку
            append_request = google_sheets_service.service.spreadsheets().values().append(
                spreadsheetId=google_sheets_service.spreadsheet_id,
                range=google_sheets_service.sheet_name,
                valueInputOption='RAW',
                insertDataOption='INSERT_ROWS',
                body={'values': test_data}
            )
            
            result = append_request.execute()
            
            return {
                "success": True,
                "message": "Google Sheets connection successful",
                "sheet_name": google_sheets_service.sheet_name,
                "spreadsheet_id": google_sheets_service.spreadsheet_id,
                "test_result": result.get('updates', {})
            }
            
        except Exception as e:
            return {
                "success": False,
                "message": f"Google Sheets write test failed: {str(e)}"
            }
            
    except Exception as e:
        logger.error(f"❌ Ошибка тестирования Google Sheets: {e}")
        raise HTTPException(status_code=500, detail=f"Test failed: {str(e)}")

def get_program_codes_by_name(program_name: str) -> List[str]:
    """
    Получает возможные коды программ по названию программы
    """
    if not program_name:
        return []
    
    program_name_lower = program_name.lower().strip()
    matching_codes = []
    
    # Ищем точные совпадения
    if program_name_lower in PROGRAM_MAPPING:
        matching_codes.append(PROGRAM_MAPPING[program_name_lower])
    
    # Ищем частичные совпадения в названиях
    for name, code in PROGRAM_MAPPING.items():
        if program_name_lower in name or name in program_name_lower:
            if code not in matching_codes:
                matching_codes.append(code)
    
    # Если ничего не найдено, возможно пользователь ввел код напрямую
    if not matching_codes:
        matching_codes.append(program_name_lower)
    
    return matching_codes

@router.get("/queue/export")
def export_queue_to_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Export all queue entries to Excel (.xlsx)"""
    # Get all queue entries
    queue_entries = get_all_queue_entries(db)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Queue Data"
    
    # Headers
    headers = ["ФИО", "Программы", "Номер", "Сотрудник", "Дата создания", "Статус", "Время обработки (сек)"]
    
    # Add headers to worksheet
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Style headers
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row, entry in enumerate(queue_entries, 2):
        programs = ", ".join(entry.programs) if isinstance(entry.programs, list) else entry.programs
        created_at = entry.created_at.strftime("%Y-%m-%d %H:%M:%S") if entry.created_at else "-"
        
        data_row = [
            entry.full_name,
            programs,
            entry.queue_number,
            entry.assigned_employee_name or "-",
            created_at,
            entry.status.value,
            entry.processing_time or "-"
        ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return file as response
    headers = {
        'Content-Disposition': 'attachment; filename="queue_data.xlsx"'
    }
    return StreamingResponse(
        io.BytesIO(output.getvalue()), 
        headers=headers, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/queue/reset-numbering")
def reset_queue_numbering(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Сбросить нумерацию очереди (только для админов)"""
    try:
        # Получаем все активные заявки (не completed)
        active_entries = db.query(QueueEntry).filter(
            QueueEntry.status.in_([QueueStatus.WAITING, QueueStatus.IN_PROGRESS, QueueStatus.PAUSED])
        ).order_by(QueueEntry.created_at).all()
        
        # Архивируем и удаляем все completed заявки
        completed_entries = db.query(QueueEntry).filter(QueueEntry.status == QueueStatus.COMPLETED).all()
        
        archived_count = 0
        for entry in completed_entries:
            try:
                archive_queue_entry(db, entry, reason="manual_reset")
                db.delete(entry)
                archived_count += 1
            except Exception as e:
                continue
        
        # Перенумеровываем активные заявки начиная с 1
        renumbered_count = 0
        for i, entry in enumerate(active_entries, 1):
            entry.queue_number = i
            db.add(entry)
            renumbered_count += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Queue numbering reset successfully",
            "archived_completed": archived_count,
            "renumbered_active": renumbered_count,
            "next_number": len(active_entries) + 1
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Reset failed: {str(e)}")

@router.post("/create-admission", response_model=UserResponse)
def create_admission_staff(
    user_data: AdminUserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Create a new admission staff member (admin only)"""
    # Create a new user with admission role
    return create_user(db=db, user=user_data, role="admission")

@router.get("/employees", response_model=List[UserResponse])
def get_all_employees(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get all employees (admin only)"""
    employees = db.query(User).filter(User.role == "admission").all()
    return employees

@router.get("/queue", response_model=List[QueueResponse])
def get_all_queue_entries_api(
    status: Optional[QueueStatus] = None,
    date: Optional[str] = None,
    employee: Optional[str] = None,
    full_name: Optional[str] = None,
    program: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get all queue entries with filters (admin only)"""
    from datetime import datetime
    
    # Начинаем с базового запроса
    query = db.query(QueueEntry)
    
    # Применяем фильтры
    if status:
        query = query.filter(QueueEntry.status == status)
    
    if date:
        # Фильтруем по дате (формат YYYY-MM-DD)
        try:
            filter_date = datetime.strptime(date, "%Y-%m-%d")
            # Фильтруем записи созданные в этот день
            start_of_day = filter_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = filter_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            query = query.filter(
                and_(
                    QueueEntry.created_at >= start_of_day,
                    QueueEntry.created_at <= end_of_day
                )
            )
        except ValueError:
            # Если неправильный формат даты, игнорируем фильтр
            pass
    
    if employee:
        # Фильтруем по имени сотрудника (частичное совпадение, без учета регистра)
        query = query.filter(
            QueueEntry.assigned_employee_name.ilike(f"%{employee}%")
        )
    
    if full_name:
        # Фильтруем по ФИО абитуриента (частичное совпадение, без учета регистра)
        query = query.filter(
            QueueEntry.full_name.ilike(f"%{full_name}%")
        )
    
    if program:
        # Получаем возможные коды программ по введенному названию
        program_codes = get_program_codes_by_name(program)
        
        if program_codes:
            # Создаем условия для поиска по всем возможным кодам
            program_conditions = []
            for code in program_codes:
                # Ищем код в JSON массиве программ
                program_conditions.append(
                    text("programs::text ILIKE :code").params(code=f'%"{code}"%')
                )
            
            # Объединяем условия через OR
            if len(program_conditions) == 1:
                query = query.filter(program_conditions[0])
            else:
                query = query.filter(or_(*program_conditions))
    
    return query.all()

@router.delete("/employees/{user_id}")
def delete_employee(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Delete employee (admin only)"""
    employee = db.query(User).filter(User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    db.delete(employee)
    db.commit()
    return {"detail": "Employee deleted successfully"}

@router.put("/employees/{user_id}", response_model=UserResponse)
def update_employee(
    user_id: str,
    user_data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Update employee data (admin only)"""
    employee = db.query(User).filter(User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    for key, value in user_data.dict(exclude_unset=True).items():
        setattr(employee, key, value)
    
    db.commit()
    db.refresh(employee)
    return employee

# === НОВЫЕ ФУНКЦИИ ДЛЯ УДАЛЕНИЯ ЗАПИСЕЙ ===

@router.delete("/queue/{queue_id}")
def delete_queue_entry_admin(
    queue_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Удалить заявку из очереди (админ) с синхронизацией"""
    try:
        # Ищем запись в основной таблице
        queue_entry = db.query(QueueEntry).filter(QueueEntry.id == queue_id).first()
        
        if queue_entry:
            # Удаляем из основной таблицы
            db.delete(queue_entry)
            logger.info(f"🗑️ Удалена запись {queue_id} из основной очереди")
        
        # Ищем запись в архиве
        archived_entry = db.query(ArchivedQueueEntry).filter(
            or_(
                ArchivedQueueEntry.id == queue_id,
                ArchivedQueueEntry.original_id == queue_id
            )
        ).first()
        
        if archived_entry:
            # 🔥 ВАЖНО: Удаляем из архива - это автоматически запустит синхронизацию с Google Sheets
            db.delete(archived_entry)
            logger.info(f"🗑️ Удалена запись {queue_id} из архива")
        
        if not queue_entry and not archived_entry:
            raise HTTPException(status_code=404, detail="Queue entry not found")
        
        db.commit()
        
        return {
            "success": True,
            "message": "Queue entry deleted successfully",
            "deleted_from_queue": bool(queue_entry),
            "deleted_from_archive": bool(archived_entry)
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Ошибка удаления заявки {queue_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting queue entry: {str(e)}")

@router.post("/queue/bulk-delete")
def bulk_delete_queue_entries(
    entry_ids: List[str],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Массовое удаление заявок с синхронизацией"""
    try:
        deleted_queue = 0
        deleted_archive = 0
        
        for queue_id in entry_ids:
            # Удаляем из основной таблицы
            queue_entry = db.query(QueueEntry).filter(QueueEntry.id == queue_id).first()
            if queue_entry:
                db.delete(queue_entry)
                deleted_queue += 1
            
            # Удаляем из архива (автоматически синхронизируется с Google Sheets)
            archived_entry = db.query(ArchivedQueueEntry).filter(
                or_(
                    ArchivedQueueEntry.id == queue_id,
                    ArchivedQueueEntry.original_id == queue_id
                )
            ).first()
            if archived_entry:
                db.delete(archived_entry)
                deleted_archive += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Bulk delete completed",
            "deleted_from_queue": deleted_queue,
            "deleted_from_archive": deleted_archive,
            "total_processed": len(entry_ids)
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Ошибка массового удаления: {e}")
        raise HTTPException(status_code=500, detail=f"Error bulk deleting: {str(e)}")

@router.post("/archive/cleanup")
def cleanup_archive(
    days_old: int = 30,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Очистка архива (удаление старых записей) с синхронизацией"""
    try:
        from datetime import datetime, timedelta
        
        cutoff_date = datetime.utcnow() - timedelta(days=days_old)
        
        # Формируем запрос для поиска старых записей
        query = db.query(ArchivedQueueEntry).filter(
            ArchivedQueueEntry.archived_at < cutoff_date
        )
        
        # Добавляем фильтр по статусу если указан
        if status_filter:
            query = query.filter(ArchivedQueueEntry.status == status_filter)
        
        # Получаем записи для подсчета
        old_entries = query.all()
        entries_count = len(old_entries)
        
        if entries_count == 0:
            return {
                "success": True,
                "message": "No entries found for cleanup",
                "deleted_count": 0,
                "cutoff_date": cutoff_date.isoformat()
            }
        
        # Удаляем найденные записи (каждая автоматически удалится из Google Sheets)
        for entry in old_entries:
            db.delete(entry)
            logger.info(f"🗑️ Удаляем старую запись {entry.id} из архива (дата: {entry.archived_at})")
        
        db.commit()
        
        logger.info(f"✅ Очистка архива завершена: удалено {entries_count} записей старше {days_old} дней")
        
        return {
            "success": True,
            "message": f"Archive cleanup completed",
            "deleted_count": entries_count,
            "cutoff_date": cutoff_date.isoformat(),
            "days_old": days_old,
            "status_filter": status_filter
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Ошибка очистки архива: {e}")
        raise HTTPException(status_code=500, detail=f"Error cleaning up archive: {str(e)}")

@router.delete("/archive/{entry_id}")
def delete_archive_entry(
    entry_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Удалить конкретную запись из архива с синхронизацией"""
    try:
        # Ищем запись в архиве
        archived_entry = db.query(ArchivedQueueEntry).filter(
            or_(
                ArchivedQueueEntry.id == entry_id,
                ArchivedQueueEntry.original_id == entry_id
            )
        ).first()
        
        if not archived_entry:
            raise HTTPException(status_code=404, detail="Archive entry not found")
        
        # Удаляем запись (автоматически синхронизируется с Google Sheets)
        db.delete(archived_entry)
        db.commit()
        
        logger.info(f"🗑️ Удалена запись {entry_id} из архива")
        
        return {
            "success": True,
            "message": "Archive entry deleted successfully",
            "entry_id": entry_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Ошибка удаления записи из архива {entry_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting archive entry: {str(e)}")

@router.post("/archive/bulk-delete")
def bulk_delete_archive_entries(
    entry_ids: List[str],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Массовое удаление записей из архива с синхронизацией"""
    try:
        deleted_count = 0
        not_found_count = 0
        
        for entry_id in entry_ids:
            # Ищем запись в архиве
            archived_entry = db.query(ArchivedQueueEntry).filter(
                or_(
                    ArchivedQueueEntry.id == entry_id,
                    ArchivedQueueEntry.original_id == entry_id
                )
            ).first()
            
            if archived_entry:
                db.delete(archived_entry)
                deleted_count += 1
                logger.info(f"🗑️ Удалена запись {entry_id} из архива")
            else:
                not_found_count += 1
                logger.warning(f"⚠️ Запись {entry_id} не найдена в архиве")
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Bulk archive delete completed",
            "deleted_count": deleted_count,
            "not_found_count": not_found_count,
            "total_processed": len(entry_ids)
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Ошибка массового удаления из архива: {e}")
        raise HTTPException(status_code=500, detail=f"Error bulk deleting from archive: {str(e)}")

@router.get("/archive/cleanup/preview")
def preview_archive_cleanup(
    days_old: int = 30,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Предварительный просмотр записей для очистки архива"""
    try:
        from datetime import datetime, timedelta
        
        cutoff_date = datetime.utcnow() - timedelta(days=days_old)
        
        # Формируем запрос для поиска старых записей
        query = db.query(ArchivedQueueEntry).filter(
            ArchivedQueueEntry.archived_at < cutoff_date
        )
        
        # Добавляем фильтр по статусу если указан
        if status_filter:
            query = query.filter(ArchivedQueueEntry.status == status_filter)
        
        # Получаем записи
        old_entries = query.limit(100).all()  # Ограничиваем для предварительного просмотра
        total_count = query.count()
        
        preview_entries = []
        for entry in old_entries:
            preview_entries.append({
                "id": entry.id,
                "original_id": entry.original_id,
                "full_name": entry.full_name,
                "status": entry.status.value if entry.status else None,
                "archived_at": entry.archived_at.isoformat() if entry.archived_at else None,
                "archive_reason": entry.archive_reason
            })
        
        return {
            "success": True,
            "total_entries_to_delete": total_count,
            "preview_entries": preview_entries,
            "cutoff_date": cutoff_date.isoformat(),
            "days_old": days_old,
            "status_filter": status_filter,
            "preview_limit": len(preview_entries)
        }
        
    except Exception as e:
        logger.error(f"❌ Ошибка предварительного просмотра очистки: {e}")
        raise HTTPException(status_code=500, detail=f"Error previewing cleanup: {str(e)}")

# === РОУТЫ ДЛЯ УПРАВЛЕНИЯ ВИДЕО ===

@router.get("/video-settings", response_model=VideoSettingsResponse)
def get_video_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get current video settings (admin only)"""
    settings = db.query(VideoSettings).first()
    if not settings:
        # Создаем запись если её нет
        settings = VideoSettings(youtube_url="", is_enabled=False)
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings

@router.put("/video-settings", response_model=VideoSettingsResponse)
def update_video_settings(
    video_data: VideoSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Update video settings (admin only)"""
    settings = db.query(VideoSettings).first()
    if not settings:
        # Создаем новую запись если её нет
        settings = VideoSettings()
        db.add(settings)
    
    # Обновляем поля
    for key, value in video_data.dict(exclude_unset=True).items():
        setattr(settings, key, value)
    
    db.commit()
    db.refresh(settings)
    return settings